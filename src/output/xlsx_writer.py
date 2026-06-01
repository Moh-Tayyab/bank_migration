from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from ..models import MigrationResult


class XLSXWriter:
    def write(self, result: MigrationResult, output_path: str):
        wb = Workbook()
        ws = wb.active
        ws.title = "Migration Report"
        ws.append(["Migration Report"])
        ws.merge_cells("A1:E1")
        ws["A1"].font = Font(bold=True, size=14)
        ws.append(["Success", "Total Records", "Processed", "Failed", "Error"])
        ws.append([
            "Yes" if result.success else "No",
            result.total_records,
            result.processed,
            result.failed,
            result.error or "",
        ])
        if result.records:
            ws.append([])
            title_row = ws.max_row
            ws.append(["Migrated Records"])
            ws[f"A{title_row + 1}"].font = Font(bold=True, size=12)
            fieldnames = list(result.records[0].keys())
            ws.append(fieldnames)
            header_fill = PatternFill(start_color="2196F3", end_color="2196F3", fill_type="solid")
            header_row = title_row + 2
            for cell in ws[header_row]:
                cell.fill = header_fill
                cell.font = Font(bold=True, color="FFFFFF")
            for record in result.records:
                ws.append([str(record.get(k, "")) for k in fieldnames])
        ws.append([])
        trail_start = ws.max_row + 1
        ws.append(["Audit Trail"])
        ws[f"A{trail_start}"].font = Font(bold=True, size=12)
        ws.append(["Event", "Record ID", "Bank Pair", "Details", "Timestamp"])
        header_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
        header_row = trail_start + 1
        for cell in ws[header_row]:
            cell.fill = header_fill
            cell.font = Font(bold=True, color="FFFFFF")
        for entry in result.audit_trail:
            ws.append([
                entry.event.value,
                entry.record_id,
                entry.bank_pair,
                entry.details,
                str(entry.timestamp),
            ])
        wb.save(output_path)