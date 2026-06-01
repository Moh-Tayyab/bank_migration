"""
Tests for Output Writers — JSON, CSV, DOCX, XLSX, HTML output generation.
"""

import csv
import json
import os

import pytest

from src.models import AuditEntry, AuditEvent, MigrationResult
from src.output import CSVWriter, DOCXWriter, HTMLWriter, JSONWriter, XLSXWriter, get_writer

# ===========================================================================
# Helper to create MigrationResult
# ===========================================================================


def make_migration_result(success=True, total=10, processed=8, failed=2, error=None):
    """Create a MigrationResult with optional audit trail entries."""
    entries = [
        AuditEntry(
            event=AuditEvent.VALIDATION,
            record_id="REC-001",
            bank_pair="A->B",
            details="Validated",
        ),
        AuditEntry(
            event=AuditEvent.MAPPING,
            record_id="REC-001",
            bank_pair="A->B",
            details="Mapped",
        ),
    ]
    return MigrationResult(
        success=success,
        total_records=total,
        processed=processed,
        failed=failed,
        audit_trail=entries,
        error=error,
    )


# ===========================================================================
# get_writer() Factory Tests
# ===========================================================================


class TestGetWriter:
    """Test the writer factory function."""

    def test_get_json_writer(self):
        writer = get_writer("json")
        assert isinstance(writer, JSONWriter)

    def test_get_csv_writer(self):
        writer = get_writer("csv")
        assert isinstance(writer, CSVWriter)

    def test_get_docx_writer(self):
        writer = get_writer("docx")
        assert isinstance(writer, DOCXWriter)

    def test_get_xlsx_writer(self):
        writer = get_writer("xlsx")
        assert isinstance(writer, XLSXWriter)

    def test_get_html_writer(self):
        writer = get_writer("html")
        assert isinstance(writer, HTMLWriter)

    def test_get_writer_case_insensitive(self):
        writer = get_writer("JSON")
        assert isinstance(writer, JSONWriter)

    def test_get_writer_unsupported_raises(self):
        with pytest.raises(ValueError, match="Unsupported output format"):
            get_writer("parquet")

    def test_get_writer_empty_raises(self):
        with pytest.raises(ValueError, match="Unsupported output format"):
            get_writer("")


# ===========================================================================
# JSON Writer Tests
# ===========================================================================


class TestJSONWriter:
    """Test JSON output writer."""

    def test_write_creates_file(self, tmp_dir):
        writer = JSONWriter()
        result = make_migration_result()
        path = os.path.join(tmp_dir, "output.json")
        writer.write(result, path)
        assert os.path.exists(path)

    def test_write_valid_json(self, tmp_dir):
        writer = JSONWriter()
        result = make_migration_result()
        path = os.path.join(tmp_dir, "output.json")
        writer.write(result, path)
        with open(path) as f:
            data = json.load(f)
        assert "migration_result" in data
        assert "audit_trail" in data

    def test_write_content(self, tmp_dir):
        writer = JSONWriter()
        result = make_migration_result(success=True, total=100, processed=95, failed=5)
        path = os.path.join(tmp_dir, "output.json")
        writer.write(result, path)
        with open(path) as f:
            data = json.load(f)
        mr = data["migration_result"]
        assert mr["success"] is True
        assert mr["total_records"] == 100
        assert mr["processed"] == 95
        assert mr["failed"] == 5

    def test_write_audit_trail(self, tmp_dir):
        writer = JSONWriter()
        result = make_migration_result()
        path = os.path.join(tmp_dir, "output.json")
        writer.write(result, path)
        with open(path) as f:
            data = json.load(f)
        assert len(data["audit_trail"]) == 2


# ===========================================================================
# CSV Writer Tests
# ===========================================================================


class TestCSVWriter:
    """Test CSV output writer."""

    def test_write_creates_file(self, tmp_dir):
        writer = CSVWriter()
        result = make_migration_result()
        path = os.path.join(tmp_dir, "output.csv")
        writer.write(result, path)
        assert os.path.exists(path)

    def test_write_valid_csv(self, tmp_dir):
        writer = CSVWriter()
        result = make_migration_result()
        path = os.path.join(tmp_dir, "output.csv")
        writer.write(result, path)
        with open(path) as f:
            reader = csv.DictReader(f)
            rows = list(reader)
        assert len(rows) == 2
        assert "event" in rows[0]

    def test_write_empty_audit_trail(self, tmp_dir):
        """CSV writer with no audit trail should write summary row."""
        writer = CSVWriter()
        result = MigrationResult(
            success=True,
            total_records=0,
            processed=0,
            failed=0,
            audit_trail=[],
        )
        path = os.path.join(tmp_dir, "output.csv")
        writer.write(result, path)
        with open(path) as f:
            reader = csv.DictReader(f)
            rows = list(reader)
        assert len(rows) == 1
        assert "success" in rows[0]


# ===========================================================================
# DOCX Writer Tests
# ===========================================================================


class TestDOCXWriter:
    """Test DOCX output writer."""

    def test_write_creates_file(self, tmp_dir):
        writer = DOCXWriter()
        result = make_migration_result()
        path = os.path.join(tmp_dir, "output.docx")
        writer.write(result, path)
        assert os.path.exists(path)

    def test_write_non_empty_file(self, tmp_dir):
        writer = DOCXWriter()
        result = make_migration_result()
        path = os.path.join(tmp_dir, "output.docx")
        writer.write(result, path)
        assert os.path.getsize(path) > 0

    def test_write_failed_migration(self, tmp_dir):
        writer = DOCXWriter()
        result = make_migration_result(success=False, error="Validation failed")
        path = os.path.join(tmp_dir, "output.docx")
        writer.write(result, path)
        assert os.path.exists(path)


# ===========================================================================
# XLSX Writer Tests
# ===========================================================================


class TestXLSXWriter:
    """Test XLSX output writer."""

    def test_write_creates_file(self, tmp_dir):
        writer = XLSXWriter()
        result = make_migration_result()
        path = os.path.join(tmp_dir, "output.xlsx")
        writer.write(result, path)
        assert os.path.exists(path)

    def test_write_non_empty_file(self, tmp_dir):
        writer = XLSXWriter()
        result = make_migration_result()
        path = os.path.join(tmp_dir, "output.xlsx")
        writer.write(result, path)
        assert os.path.getsize(path) > 0

    def test_write_contains_data(self, tmp_dir):
        from openpyxl import load_workbook

        writer = XLSXWriter()
        result = make_migration_result(success=True, total=50, processed=48, failed=2)
        path = os.path.join(tmp_dir, "output.xlsx")
        writer.write(result, path)
        wb = load_workbook(path)
        ws = wb.active
        # Should have migration report data
        assert ws["A1"].value == "Migration Report"
        wb.close()


# ===========================================================================
# HTML Writer Tests
# ===========================================================================


class TestHTMLWriter:
    """Test HTML output writer."""

    def test_write_creates_file(self, tmp_dir):
        writer = HTMLWriter()
        result = make_migration_result()
        path = os.path.join(tmp_dir, "output.html")
        writer.write(result, path)
        assert os.path.exists(path)

    def test_write_valid_html(self, tmp_dir):
        writer = HTMLWriter()
        result = make_migration_result()
        path = os.path.join(tmp_dir, "output.html")
        writer.write(result, path)
        with open(path) as f:
            content = f.read()
        assert "<html>" in content
        assert "</html>" in content
        assert "Migration Report" in content

    def test_write_contains_audit_trail(self, tmp_dir):
        writer = HTMLWriter()
        result = make_migration_result()
        path = os.path.join(tmp_dir, "output.html")
        writer.write(result, path)
        with open(path) as f:
            content = f.read()
        assert "VALIDATION" in content
        assert "MAPPING" in content

    def test_write_contains_summary(self, tmp_dir):
        writer = HTMLWriter()
        result = make_migration_result(total=100, processed=95, failed=5)
        path = os.path.join(tmp_dir, "output.html")
        writer.write(result, path)
        with open(path) as f:
            content = f.read()
        assert "100" in content
        assert "95" in content
