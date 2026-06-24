import hmac
import json
import logging
import mimetypes
import os
import re
import uuid
from contextlib import asynccontextmanager
from datetime import datetime
from typing import Optional

import uvicorn
from fastapi import Depends, FastAPI, File, Form, HTTPException, Request, Security, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
from fastapi.security import APIKeyHeader
from slowapi import Limiter, _rate_limit_exceeded_handler
from slowapi.errors import RateLimitExceeded
from slowapi.util import get_remote_address

from src.config import settings
from src.infrastructure.retention import DataRetentionPolicy
from src.production import PipelineOrchestrator

logger = logging.getLogger(__name__)

API_KEY_NAME = "X-API-Key"
_api_key_header = APIKeyHeader(name=API_KEY_NAME, auto_error=False)

limiter = Limiter(key_func=get_remote_address)


async def verify_api_key(api_key: str = Security(_api_key_header)):
    expected = os.getenv("API_KEY", "")
    if not expected:
        env = os.getenv("ENVIRONMENT", "development")
        if env == "production":
            logger.error("API_KEY not set in production — refusing unauthenticated requests")
            raise HTTPException(status_code=500, detail="Authentication not configured")
        logger.warning("API_KEY not set — authentication disabled (development only)")
        return None
    if not api_key or not hmac.compare_digest(api_key, expected):
        raise HTTPException(status_code=401, detail="Invalid or missing API key")
    return api_key


def _get_celery_tasks():
    try:
        from src.infrastructure.tasks import run_data_migration_task, run_full_migration_task, run_multi_migration_task

        return run_full_migration_task, run_data_migration_task, run_multi_migration_task
    except Exception:
        logger.debug("Celery tasks not available, falling back to synchronous processing")
        return None, None, None


@asynccontextmanager
async def lifespan(app: FastAPI):
    if settings.cleanup_on_startup:
        try:
            policy = DataRetentionPolicy()
            report = policy.run_all()
            if report.total_deleted > 0:
                logger.info(
                    "Startup cleanup: deleted %d uploads, %d output, %d audit, %d canonical",
                    report.uploads_deleted,
                    report.output_deleted,
                    report.audit_deleted,
                    report.canonical_deleted,
                )
        except Exception as e:
            logger.warning("Startup cleanup failed (non-fatal): %s", e)
    yield


app = FastAPI(
    title="UN Wallet Multi-Bank Data Migration API",
    description="Production-Grade Interbank ETL Platform",
    version="1.0.0",
    lifespan=lifespan,
    docs_url=None if os.getenv("ENVIRONMENT") == "production" else "/docs",
    redoc_url=None if os.getenv("ENVIRONMENT") == "production" else "/redoc",
    openapi_url=None if os.getenv("ENVIRONMENT") == "production" else "/openapi.json",
)
app.state.limiter = limiter
app.add_exception_handler(RateLimitExceeded, _rate_limit_exceeded_handler)

cors_origins = os.getenv("CORS_ORIGINS", "http://localhost:3000").split(",")
app.add_middleware(
    CORSMiddleware,
    allow_origins=cors_origins,
    allow_credentials=True,
    allow_methods=["GET", "POST"],
    allow_headers=["X-API-Key", "Content-Type", "Accept"],
)


@app.middleware("http")
async def security_headers_middleware(request: Request, call_next):
    response = await call_next(request)
    response.headers["X-Content-Type-Options"] = "nosniff"
    response.headers["X-Frame-Options"] = "DENY"
    response.headers["Referrer-Policy"] = "strict-origin-when-cross-origin"
    response.headers["Permissions-Policy"] = "camera=(), microphone=(), geolocation=()"
    return response


orchestrator = PipelineOrchestrator()
_schema_ai = None
_anomaly_ai = None

# Directory to store previewed files for reuse
_preview_store_dir = settings.upload_dir / "preview_store"
_preview_store_dir.mkdir(parents=True, exist_ok=True)


def _cleanup_old_previews():
    """Remove preview files older than preview_store_ttl_hours."""
    import time

    try:
        now = time.time()
        ttl_seconds = settings.preview_store_ttl_hours * 3600
        removed = 0
        for filepath in _preview_store_dir.iterdir():
            if filepath.is_file():
                if now - filepath.stat().st_mtime > ttl_seconds:
                    try:
                        filepath.unlink()
                        removed += 1
                    except OSError:
                        pass
        if removed > 0:
            logger.info("Cleaned up %d old preview files", removed)
    except Exception as e:
        logger.warning("Preview cleanup failed: %s", e)


def get_schema_ai():
    global _schema_ai
    if _schema_ai is None:
        from src.ai.schema_agent import SchemaIntelligenceAgent

        _schema_ai = SchemaIntelligenceAgent()
    return _schema_ai


def get_anomaly_ai():
    global _anomaly_ai
    if _anomaly_ai is None:
        from src.ai.anomaly_agent import AnomalyDetectionAgent

        _anomaly_ai = AnomalyDetectionAgent()
    return _anomaly_ai


@app.get("/api/health")
async def health():
    return {"status": "healthy", "timestamp": datetime.utcnow().isoformat()}


@app.post("/api/migrate/upload")
@limiter.limit("10/minute")
async def migrate_upload(
    request: Request,
    _auth=Depends(verify_api_key),
    file: Optional[UploadFile] = File(None),  # Optional when file_id is provided
    source_bank: str = Form(...),
    target_banks: str = Form("[]"),
    output_format: Optional[str] = Form("json"),
    file_id: Optional[str] = Form(None),  # Reuse previously uploaded file
):
    import json as _json

    banks = _json.loads(target_banks) if isinstance(target_banks, str) else target_banks

    # Determine file source: uploaded file OR stored file_id
    filepath = None
    is_temporary = False

    if file_id:
        # Reuse stored file from preview
        matching_files = list(_preview_store_dir.glob(f"{file_id}_*"))
        if not matching_files:
            raise HTTPException(status_code=404, detail="Preview file not found or expired. Please re-upload the file.")
        filepath = str(matching_files[0])
        logger.info(f"Reusing stored preview file: {filepath}")
    elif file:
        # New file upload
        os.makedirs(settings.upload_dir, exist_ok=True)
        safe_filename = os.path.basename(file.filename or "upload")
        new_file_id = str(uuid.uuid4())
        filepath = os.path.join(settings.upload_dir, f"{new_file_id}_{safe_filename}")
        is_temporary = True
        max_bytes = settings.max_file_size_mb * 1024 * 1024
        try:
            with open(filepath, "wb") as f:
                total = 0
                while chunk := await file.read(1024 * 1024):
                    total += len(chunk)
                    if total > max_bytes:
                        raise HTTPException(status_code=413, detail=f"File exceeds {settings.max_file_size_mb}MB limit")
                    f.write(chunk)
        except BaseException:
            if os.path.exists(filepath):
                os.remove(filepath)
            raise
    else:
        raise HTTPException(status_code=400, detail="Either file or file_id must be provided")

    celery_dispatched = False
    try:
        if len(banks) == 1:
            run_full, _, _ = _get_celery_tasks()
            if run_full:
                try:
                    task = run_full.delay(filepath, source_bank, banks[0], output_format)
                    celery_dispatched = True
                    task_id = task.id
                    return {
                        "task_id": task_id,
                        "status": "queued",
                        "message": (
                            f"Migration to {len(banks)} target bank(s) started in "
                            f"background. Use /api/status/{task_id} to check progress."
                        ),
                        "file_id": file_id,
                    }
                except Exception:
                    logger.debug("Celery task dispatch failed, running synchronously")
            result = orchestrator.migrate_file(filepath, source_bank, banks[0], output_format)
            return json.loads(result.model_dump_json())
        else:
            _, _, run_multi = _get_celery_tasks()
            if run_multi:
                try:
                    task = run_multi.delay(filepath, source_bank, banks, output_format)
                    celery_dispatched = True
                    task_id = task.id
                    return {
                        "task_id": task_id,
                        "status": "queued",
                        "message": (
                            f"Migration to {len(banks)} target bank(s) started in "
                            f"background. Use /api/status/{task_id} to check progress."
                        ),
                        "file_id": file_id,
                    }
                except Exception:
                    logger.debug("Celery multi-task dispatch failed, running synchronously")
            result = orchestrator.migrate_file_multi(filepath, source_bank, banks, output_format)
            return json.loads(result.model_dump_json())
    except HTTPException:
        raise
    except Exception:
        logger.exception("Migration failed")
        raise HTTPException(status_code=500, detail="Migration failed. Check server logs for details.")
    finally:
        # Only clean up if it was a temporary upload (not from preview_store)
        if is_temporary and not celery_dispatched and os.path.exists(filepath):
            try:
                os.remove(filepath)
            except OSError:
                pass


@app.post("/api/migrate/data")
@limiter.limit("10/minute")
async def migrate_data(
    request: Request,
    _auth=Depends(verify_api_key),
):
    import json as _json

    body = await request.json()
    records = body.get("records", [])
    source_bank = body.get("source_bank", "")
    target_banks_raw = body.get("target_banks", [])
    output_format = body.get("output_format", "json")
    banks = target_banks_raw if isinstance(target_banks_raw, list) else _json.loads(target_banks_raw)

    if len(banks) == 1:
        _, run_data, _ = _get_celery_tasks()
        if run_data:
            try:
                task = run_data.delay(records, source_bank, banks[0], output_format)
                return {
                    "task_id": task.id,
                    "status": "queued",
                    "message": f"Migration to {len(banks)} target bank(s) started in background.",
                }
            except Exception:
                logger.debug("Celery data task dispatch failed, running synchronously")
        result = orchestrator.migrate_data(records, source_bank, banks[0], output_format)
        return json.loads(result.model_dump_json())
    else:
        _, _, run_multi = _get_celery_tasks()
        if run_multi:
            try:
                task = run_multi.delay(records, source_bank, banks, output_format)
                return {
                    "task_id": task.id,
                    "status": "queued",
                    "message": f"Migration to {len(banks)} target bank(s) started in background.",
                }
            except Exception:
                logger.debug("Celery multi-data task dispatch failed, running synchronously")
        result = orchestrator.migrate_data_multi(records, source_bank, banks, output_format)
        return json.loads(result.model_dump_json())


@app.get("/api/status/{task_id}")
@limiter.limit("30/minute")
async def get_task_status(request: Request, task_id: str, _auth=Depends(verify_api_key)):
    try:
        from src.infrastructure.celery_app import app as celery_app

        task_result = celery_app.AsyncResult(task_id)
        response = {
            "task_id": task_id,
            "status": task_result.status,
            "result": task_result.result if task_result.ready() else None,
        }
        return response
    except Exception:
        logger.debug("Celery status check failed")
        return {"task_id": task_id, "status": "unavailable", "result": None}


@app.get("/api/banks")
@limiter.limit("30/minute")
async def list_banks(request: Request, _auth=Depends(verify_api_key)):
    return {"banks": orchestrator.get_banks()}


@app.get("/api/schema/{source_bank}/{target_bank}")
@limiter.limit("30/minute")
async def get_schema_mapping(request: Request, source_bank: str, target_bank: str, _auth=Depends(verify_api_key)):
    mappings = orchestrator.get_schema_mapping(source_bank, target_bank)
    return {"source_bank": source_bank, "target_bank": target_bank, "mappings": mappings}


@app.post("/api/schema/auto-map")
@limiter.limit("10/minute")
async def auto_map_columns(
    request: Request,
    _auth=Depends(verify_api_key),
    file: UploadFile = File(...),
    target_bank: Optional[str] = Form(None),
):
    from src.schema_mapper import auto_generate_mappings

    os.makedirs(settings.upload_dir, exist_ok=True)
    safe_filename = os.path.basename(file.filename or "upload")
    file_id = str(uuid.uuid4())
    filepath = os.path.join(settings.upload_dir, f"{file_id}_{safe_filename}")

    max_bytes = settings.max_file_size_mb * 1024 * 1024
    try:
        with open(filepath, "wb") as f:
            total = 0
            while chunk := await file.read(1024 * 1024):
                total += len(chunk)
                if total > max_bytes:
                    raise HTTPException(status_code=413, detail=f"File exceeds {settings.max_file_size_mb}MB limit")
                f.write(chunk)
    except BaseException:
        if os.path.exists(filepath):
            os.remove(filepath)
        raise

    try:
        detected, records = orchestrator.preview_file(filepath, row_limit=5)
        columns = list(records[0].keys()) if records else []

        if not target_bank:
            target_bank = orchestrator.detect_target_bank(columns)
        if not target_bank:
            raise HTTPException(
                status_code=400,
                detail="target_bank not provided and could not be auto-detected from the file",
            )

        mappings = auto_generate_mappings(columns, target_bank)
        mapping_list = [{"source": m.source_field, "target": m.target_field} for m in mappings]

        matched_count = len(mapping_list)
        total_source = len(columns)

        return {
            "filename": file.filename,
            "format": detected,
            "source_columns": columns,
            "target_bank": target_bank,
            "mappings": mapping_list,
            "matched": matched_count,
            "total_source_columns": total_source,
            "preview_rows": records[:3],
            "file_id": file_id,
        }
    except Exception:
        if os.path.exists(filepath):
            os.remove(filepath)
        raise


@app.post("/api/schema/upload-target")
@limiter.limit("10/minute")
async def upload_target_file(
    request: Request,
    _auth=Depends(verify_api_key),
    file: UploadFile = File(...),
):
    """Parse uploaded target file and return its columns for schema inference."""
    from src.schema_mapper import generate_custom_mappings

    os.makedirs(settings.upload_dir, exist_ok=True)
    safe_filename = os.path.basename(file.filename or "upload")
    file_id = str(uuid.uuid4())
    filepath = os.path.join(settings.upload_dir, f"{file_id}_{safe_filename}")

    max_bytes = settings.max_file_size_mb * 1024 * 1024
    try:
        with open(filepath, "wb") as f:
            total = 0
            while chunk := await file.read(1024 * 1024):
                total += len(chunk)
                if total > max_bytes:
                    raise HTTPException(status_code=413, detail=f"File exceeds {settings.max_file_size_mb}MB limit")
                f.write(chunk)
    except BaseException:
        if os.path.exists(filepath):
            os.remove(filepath)
        raise

    try:
        detected, records = orchestrator.preview_file(filepath, row_limit=5)
        target_columns = list(records[0].keys()) if records else []
        sample_values = records[0] if records else {}

        return {
            "file_id": file_id,
            "filename": file.filename,
            "format": detected,
            "columns": target_columns,
            "sample_values": sample_values,
            "row_count": len(records),
        }
    except Exception:
        if os.path.exists(filepath):
            os.remove(filepath)
        raise


@app.post("/api/schema/auto-map-custom")
@limiter.limit("10/minute")
async def auto_map_custom_columns(
    request: Request,
    _auth=Depends(verify_api_key),
    source_columns: str = Form(...),
    target_columns: str = Form(...),
):
    """Generate auto-mappings between source file columns and target file columns."""
    from src.schema_mapper import generate_custom_mappings

    import json as _json

    src_cols = _json.loads(source_columns)
    tgt_cols = _json.loads(target_columns)

    mappings = generate_custom_mappings(src_cols, tgt_cols)
    mapping_list = [{"source": m.source_field, "target": m.target_field} for m in mappings]

    return {
        "source_columns": src_cols,
        "target_columns": tgt_cols,
        "mappings": mapping_list,
        "matched": len(mapping_list),
        "unmatched_source": [c for c in src_cols if c not in [m["source"] for m in mapping_list]],
        "unmatched_target": [c for c in tgt_cols if c not in [m["target"] for m in mapping_list]],
    }


@app.post("/api/migrate/upload-custom")
@limiter.limit("10/minute")
async def migrate_custom_upload(
    request: Request,
    _auth=Depends(verify_api_key),
    source_file: UploadFile = File(...),
    target_file: UploadFile = File(...),
    output_format: Optional[str] = Form("json"),
    mappings: Optional[str] = Form(None),
):
    """Migrate data from source file to match target file's schema."""
    from src.schema_mapper import generate_custom_mappings, register_custom_target

    import json as _json

    os.makedirs(settings.upload_dir, exist_ok=True)

    # Save source file
    src_id = str(uuid.uuid4())
    src_safe = os.path.basename(source_file.filename or "source")
    src_path = os.path.join(settings.upload_dir, f"{src_id}_{src_safe}")

    max_bytes = settings.max_file_size_mb * 1024 * 1024
    try:
        with open(src_path, "wb") as f:
            total = 0
            while chunk := await source_file.read(1024 * 1024):
                total += len(chunk)
                if total > max_bytes:
                    raise HTTPException(status_code=413, detail=f"Source file exceeds {settings.max_file_size_mb}MB limit")
                f.write(chunk)
    except BaseException:
        if os.path.exists(src_path):
            os.remove(src_path)
        raise

    # Save target file
    tgt_id = str(uuid.uuid4())
    tgt_safe = os.path.basename(target_file.filename or "target")
    tgt_path = os.path.join(settings.upload_dir, f"{tgt_id}_{tgt_safe}")

    try:
        with open(tgt_path, "wb") as f:
            total = 0
            while chunk := await target_file.read(1024 * 1024):
                total += len(chunk)
                if total > max_bytes:
                    raise HTTPException(status_code=413, detail=f"Target file exceeds {settings.max_file_size_mb}MB limit")
                f.write(chunk)
    except BaseException:
        if os.path.exists(tgt_path):
            os.remove(tgt_path)
        raise

    try:
        # Parse both files
        src_detected, src_records = orchestrator.preview_file(src_path, row_limit=999999)
        tgt_detected, tgt_records = orchestrator.preview_file(tgt_path, row_limit=5)

        src_columns = list(src_records[0].keys()) if src_records else []
        tgt_columns = list(tgt_records[0].keys()) if tgt_records else []

        if not src_records:
            raise HTTPException(status_code=400, detail="Source file contains no records")
        if not tgt_columns:
            raise HTTPException(status_code=400, detail="Target file contains no columns")

        # Register custom target schema
        sample_values = tgt_records[0] if tgt_records else {}
        custom_bank_name = f"custom_{tgt_id[:8]}"
        register_custom_target(
            orchestrator._registry,
            tgt_columns,
            sample_values,
            custom_bank_name,
        )

        # Use provided mappings or auto-generate
        if mappings:
            mapping_list = _json.loads(mappings)
            from src.models import MappingRule
            custom_mappings = [
                MappingRule(
                    source_field=m["source"],
                    target_field=m["target"],
                    transform=m.get("transform", ""),
                    required=m.get("required", False),
                    default=m.get("default"),
                )
                for m in mapping_list
            ]
        else:
            custom_mappings = generate_custom_mappings(src_columns, tgt_columns)

        # Run migration with custom mappings
        from src.models import MappingRule, Record
        from src.pipeline import Pipeline
        from src.stages import ValidateStage, ParseStage, MapStage, RulesStage, StoreStage, MaskStage
        from src.rules_engine import build_standard_rules
        from src.transaction_rollback import TransactionManager
        from src.audit_logger import AuditLogger

        audit = AuditLogger()
        pipeline = Pipeline(
            stages=[
                ValidateStage(),
                ParseStage(),
                MapStage(),
                RulesStage(engine=build_standard_rules()),
                StoreStage(),
                MaskStage(),
            ],
            txn=TransactionManager(),
            audit=audit,
        )

        # Inject custom mappings into the mapper
        pipeline._stages[2]._mapper._auto_cache[custom_bank_name] = custom_mappings

        # Run pipeline
        result = pipeline.run(iter(src_records), "__auto__", custom_bank_name)

        # Generate output
        if result.success and result.processed > 0:
            fmt = output_format or "json"
            from src.output import get_writer
            writer = get_writer(fmt)
            timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
            output_path = settings.output_dir / f"migration_custom_{timestamp}.{fmt}"
            settings.output_dir.mkdir(parents=True, exist_ok=True)
            writer.write(result, str(output_path))
            result.output_path = str(output_path)

        return _json.loads(result.model_dump_json())

    except HTTPException:
        raise
    except Exception as e:
        logger.exception("Custom migration failed")
        raise HTTPException(status_code=500, detail=f"Migration failed: {str(e)}")
    finally:
        # Cleanup temp files
        for path in [src_path, tgt_path]:
            if os.path.exists(path):
                try:
                    os.remove(path)
                except OSError:
                    pass


@app.get("/api/download/{filename}")
@limiter.limit("20/minute")
async def download_file(request: Request, filename: str, _auth=Depends(verify_api_key)):
    safe_name = os.path.basename(filename)
    filepath = settings.output_dir / safe_name
    if not filepath.is_file():
        raise HTTPException(status_code=404, detail="File not found")
    media_type, _ = mimetypes.guess_type(str(filepath))
    return FileResponse(
        path=str(filepath),
        media_type=media_type or "application/octet-stream",
        filename=safe_name,
    )


@app.post("/api/preview")
@limiter.limit("10/minute")
async def preview_file(
    request: Request,
    _auth=Depends(verify_api_key),
    file: UploadFile = File(...),
    row_limit: int = Form(10),
    source_bank: Optional[str] = Form(None),
):
    row_limit = min(row_limit, 100)
    _cleanup_old_previews()  # Clean up old files before storing new one

    safe_filename = os.path.basename(file.filename or "upload")
    file_id = str(uuid.uuid4())
    # Store in preview_store directory for reuse during migration
    filepath = _preview_store_dir / f"{file_id}_{safe_filename}"
    max_bytes = settings.max_file_size_mb * 1024 * 1024
    try:
        with open(filepath, "wb") as f:
            total = 0
            while chunk := await file.read(1024 * 1024):
                total += len(chunk)
                if total > max_bytes:
                    raise HTTPException(status_code=413, detail=f"File exceeds {settings.max_file_size_mb}MB limit")
                f.write(chunk)
    except BaseException:
        if os.path.exists(filepath):
            os.remove(filepath)
        raise

    try:
        detected, records = orchestrator.preview_file(str(filepath), row_limit)
        columns = list(records[0].keys()) if records else []

        # Auto-detect target bank based on column matching
        detected_target = None
        if columns:
            exclude = [source_bank] if source_bank else []
            detected_target = orchestrator.detect_target_bank(columns, exclude_banks=exclude)

        return {
            "filename": file.filename,
            "format": detected,
            "total_columns": len(columns),
            "columns": columns,
            "rows": records,
            "row_count": len(records),
            "detected_target_bank": detected_target,
            "file_id": file_id,  # Return file_id for reuse in migration
            "stored_filename": f"{file_id}_{safe_filename}",
        }
    except Exception:
        # Clean up file if preview fails
        if os.path.exists(filepath):
            os.remove(filepath)
        raise


@app.post("/api/sqlldr/generate")
@limiter.limit("10/minute")
async def generate_migrate_script(
    request: Request,
    _auth=Depends(verify_api_key),
    file: UploadFile = File(...),
    table_name: Optional[str] = Form(None),
    target_file: Optional[UploadFile] = File(None),
    mappings: Optional[str] = Form(None),
    output_format: Optional[str] = Form("csv"),
):
    """
    Generate a self-contained migration script (PowerShell for Windows,
    Python for Mac/Linux). Each script reads the embedded source data,
    applies the column mappings, and writes the output in the requested
    format. Supported formats: csv, json, html, xlsx.
    """
    import csv
    import base64

    os.makedirs(settings.upload_dir, exist_ok=True)
    os.makedirs(settings.output_dir, exist_ok=True)

    safe_filename = os.path.basename(file.filename or "data.csv")
    file_id = str(uuid.uuid4())
    filepath = os.path.join(settings.upload_dir, f"{file_id}_{safe_filename}")

    max_bytes = settings.max_file_size_mb * 1024 * 1024
    try:
        with open(filepath, "wb") as f:
            total = 0
            while chunk := await file.read(1024 * 1024):
                total += len(chunk)
                if total > max_bytes:
                    raise HTTPException(status_code=413, detail=f"File exceeds {settings.max_file_size_mb}MB limit")
                f.write(chunk)
    except BaseException:
        if os.path.exists(filepath):
            os.remove(filepath)
        raise

    # Parse mappings
    mapping_list = []
    if mappings:
        try:
            mapping_list = json.loads(mappings)
        except Exception:
            mapping_list = []

    col_map = {}
    for m in mapping_list:
        src = m.get("source", "")
        tgt = m.get("target", "")
        if src and tgt:
            col_map[src] = tgt

    try:
        # Read source file
        records = []
        columns = []
        file_ext = os.path.splitext(safe_filename)[1].lower()

        if file_ext in (".xlsx", ".xls"):
            try:
                import openpyxl
                wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
                ws = wb.active
                rows = list(ws.iter_rows(values_only=True))
                if rows:
                    columns = [str(c) if c is not None else f"COL_{i}" for i, c in enumerate(rows[0])]
                    for row in rows[1:]:
                        records.append({columns[i]: (str(v) if v is not None else None) for i, v in enumerate(row) if i < len(columns)})
                wb.close()
            except ImportError:
                raise HTTPException(status_code=400, detail="Excel support requires openpyxl")
        elif file_ext == ".json":
            with open(filepath, "r", encoding="utf-8") as f:
                data = json.load(f)
            if isinstance(data, list) and len(data) > 0 and isinstance(data[0], dict):
                columns = list(data[0].keys())
                for item in data:
                    records.append({k: (str(v) if v is not None else None) for k, v in item.items()})
        else:
            raw_bytes = open(filepath, "rb").read()
            decoded = None
            for enc in ("utf-8-sig", "utf-8", "latin-1", "cp1252", "iso-8859-1", "ascii"):
                try:
                    decoded = raw_bytes.decode(enc)
                    break
                except (UnicodeDecodeError, LookupError):
                    continue
            if decoded is None:
                decoded = raw_bytes.decode("utf-8", errors="replace")

            import io
            fio = io.StringIO(decoded)
            sniffer = csv.Sniffer()
            delimiter = ","
            try:
                delimiter = sniffer.sniff(decoded[:1024]).delimiter
            except Exception:
                delimiter = ","
            reader = csv.DictReader(fio, delimiter=delimiter)
            if reader.fieldnames:
                columns = list(reader.fieldnames)
            for row in reader:
                records.append({k: (v if v != "" else None) for k, v in row.items()})

        if not records:
            raise HTTPException(status_code=400, detail="No records found in uploaded file")

        # --- Security seam (candidate #2): mask PII BEFORE it is embedded. ---
        # Path A used to base64-embed the raw source file, leaking PII inside the
        # generated script. We now run every record through SecurityMasker (the same
        # component the real pipeline's SecurityStage uses) and re-serialize to the
        # source format, so the embedded payload is safe-by-construction for
        # csv/json/xlsx alike. See Step 3 of the architecture review.
        from src.security import SecurityMasker
        _masker = SecurityMasker()
        records = [_masker.mask(rec, str(idx)) for idx, rec in enumerate(records)]

        def _serialize_payload(rows, cols, ext):
            import io as _io, csv as _csv
            if ext == ".json":
                return json.dumps(rows, ensure_ascii=False).encode("utf-8")
            if ext in (".xlsx", ".xls"):
                import openpyxl
                wb = openpyxl.Workbook(); ws = wb.active; ws.append(list(cols))
                for rec in rows:
                    ws.append([rec.get(c) for c in cols])
                buf = _io.BytesIO(); wb.save(buf); return buf.getvalue()
            sio = _io.StringIO()
            w = _csv.writer(sio); w.writerow(list(cols))
            for rec in rows:
                w.writerow([rec.get(c) for c in cols])
            return sio.getvalue().encode("utf-8")

        # Target columns
        target_columns = []
        for col in columns:
            t = col_map.get(col, col)
            target_columns.append(t)

        # Read target file columns if provided
        target_file_columns = []
        if target_file:
            tf_path = os.path.join(settings.upload_dir, f"{file_id}_target_{target_file.filename}")
            with open(tf_path, "wb") as f:
                while chunk := await target_file.read(1024 * 1024):
                    f.write(chunk)
            tf_ext = os.path.splitext(target_file.filename or "")[1].lower()
            if tf_ext in (".xlsx", ".xls"):
                import openpyxl
                wb = openpyxl.load_workbook(tf_path, read_only=True, data_only=True)
                ws = wb.active
                rows = list(ws.iter_rows(values_only=True))
                if rows:
                    target_file_columns = [str(c) for c in rows[0]]
                wb.close()
            elif tf_ext == ".json":
                with open(tf_path, "r", encoding="utf-8") as f:
                    td = json.load(f)
                if isinstance(td, list) and len(td) > 0 and isinstance(td[0], dict):
                    target_file_columns = list(td[0].keys())
            else:
                raw = open(tf_path, "rb").read()
                dec = None
                for enc in ("utf-8-sig", "utf-8", "latin-1", "cp1252"):
                    try:
                        dec = raw.decode(enc)
                        break
                    except (UnicodeDecodeError, LookupError):
                        continue
                if dec:
                    import io
                    reader = csv.DictReader(io.StringIO(dec))
                    if reader.fieldnames:
                        target_file_columns = list(reader.fieldnames)
            os.remove(tf_path)

        # Build the Python migration script
        mapping_dict_str = json.dumps(col_map, indent=4) if col_map else "{}"
        source_file_b64 = ""
        source_ext = file_ext if file_ext in (".csv", ".json", ".xlsx", ".xls") else ".csv"

        # Embed the MASKED payload (re-serialized to the source format) so the
        # generated script never carries raw PII. The Python branch decodes by file
        # extension, so we preserve the original extension.
        source_file_b64 = base64.b64encode(_serialize_payload(records, columns, file_ext)).decode("ascii")

        source_filename_safe = safe_filename.replace('"', '\\"')

        # Determine output format from source extension
        output_ext = ".csv"

        # Build encoded values for embedding in scripts
        mapping_json = mapping_dict_str

        # For PowerShell: convert Excel/JSON to CSV text before base64 encoding
        ps_source_b64 = source_file_b64
        ps_source_ext = source_ext
        if file_ext in (".xlsx", ".xls"):
            import io as _io, csv as _csv
            csv_buffer = _io.StringIO()
            writer = _csv.writer(csv_buffer)
            writer.writerow(columns)
            for rec in records:
                writer.writerow([rec.get(c, None) for c in columns])
            ps_csv_data = csv_buffer.getvalue()
            ps_source_b64 = base64.b64encode(ps_csv_data.encode("utf-8")).decode("ascii")
            ps_source_ext = ".csv"
        elif file_ext == ".json":
            ps_csv_data = ",".join(columns) + "\n"
            for rec in records:
                ps_csv_data += ",".join(str(rec.get(c, "") or "") for c in columns) + "\n"
            ps_source_b64 = base64.b64encode(ps_csv_data.encode("utf-8")).decode("ascii")
            ps_source_ext = ".csv"

        def to_ps_array(lst):
            return '@("' + '", "'.join(lst) + '")'

        ps_target_cols = to_ps_array(target_columns)
        source_columns_json = json.dumps(columns)
        target_columns_json = json.dumps(target_columns)

        # ---- Output format handling (csv | json | html | xlsx) ----
        fmt = (output_format or "csv").strip().lower()
        if fmt not in ("csv", "json", "html", "xlsx"):
            fmt = "csv"

        # Python write block, embedded verbatim into the generated .py script.
        # These are plain strings (not f-strings) so their braces are literal.
        if fmt == "csv":
            py_write_block = '''    out = os.path.splitext(SOURCE_FILENAME)[0] + "_migrated.csv"
    with open(out, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=TARGET_COLUMNS, extrasaction="ignore")
        w.writeheader(); w.writerows(mapped)'''
        elif fmt == "json":
            py_write_block = '''    out = os.path.splitext(SOURCE_FILENAME)[0] + "_migrated.json"
    with open(out, "w", encoding="utf-8") as f:
        json.dump(mapped, f, ensure_ascii=False, indent=2)'''
        elif fmt == "html":
            py_write_block = '''    def _esc(v): return ("" if v is None else str(v)).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
    out = os.path.splitext(SOURCE_FILENAME)[0] + "_migrated.html"
    parts = ["<table><tr>" + "".join(f"<th>{_esc(c)}</th>" for c in TARGET_COLUMNS) + "</tr>"]
    for rec in mapped:
        parts.append("<tr>" + "".join(f"<td>{_esc(rec.get(c))}</td>" for c in TARGET_COLUMNS) + "</tr>")
    parts.append("</table>")
    with open(out, "w", encoding="utf-8") as f:
        f.write("<!doctype html><html><head><meta charset='utf-8'><title>migrated</title></head><body>" + "".join(parts) + "</body></html>")'''
        else:  # xlsx
            py_write_block = '''    out = os.path.splitext(SOURCE_FILENAME)[0] + "_migrated.xlsx"
    try:
        import openpyxl
    except ImportError:
        print("ERROR: xlsx output requires openpyxl. Install with: pip install openpyxl"); sys.exit(1)
    wb = openpyxl.Workbook(); ws = wb.active
    ws.append(TARGET_COLUMNS)
    for rec in mapped:
        ws.append([rec.get(c) for c in TARGET_COLUMNS])
    wb.save(out)'''

        # PowerShell write block. PowerShell can't author .xlsx without Excel,
        # so xlsx falls back to CSV here with a note (use the Python script for xlsx).
        if fmt == "csv":
            ps_write_block = '''$outputName = [System.IO.Path]::GetFileNameWithoutExtension($sourceFilename) + "_migrated.csv"
$mapped | Select-Object $targetColumns | Export-Csv -Path $outputName -NoTypeInformation -Encoding UTF8'''
        elif fmt == "json":
            ps_write_block = '''$outputName = [System.IO.Path]::GetFileNameWithoutExtension($sourceFilename) + "_migrated.json"
$mapped | ConvertTo-Json -Depth 5 | Out-File -FilePath $outputName -Encoding UTF8'''
        elif fmt == "html":
            ps_write_block = '''$outputName = [System.IO.Path]::GetFileNameWithoutExtension($sourceFilename) + "_migrated.html"
$mapped | Select-Object $targetColumns | ConvertTo-Html | Out-File -FilePath $outputName -Encoding UTF8'''
        else:  # xlsx -> CSV fallback
            ps_write_block = '''Write-Host "Note: .xlsx is not supported in PowerShell (requires Excel). Wrote CSV instead. For .xlsx output, run the Python script."
$outputName = [System.IO.Path]::GetFileNameWithoutExtension($sourceFilename) + "_migrated.csv"
$mapped | Select-Object $targetColumns | Export-Csv -Path $outputName -NoTypeInformation -Encoding UTF8'''

        # ============ POWERSHELL SCRIPT (Windows - no Python needed) ============
        ps_source_name = os.path.splitext(safe_filename)[0] + ".csv"
        ps_script = f'''# Migration Script - Generated by UN Wallet
# Source: {safe_filename} | Records: {len(records)} | Columns: {len(columns)} -> {len(target_columns)} | Output: {fmt}

$dataB64 = @"
{ps_source_b64}
"@

$mappingsJson = @'
{mapping_json}
'@

$targetColumns = {ps_target_cols}
$sourceFilename = "{ps_source_name}"

# Decode base64 source data
$bytes = [Convert]::FromBase64String($dataB64)
$decoded = [System.Text.Encoding]::UTF8.GetString($bytes)

# Parse CSV
$records = $decoded | ConvertFrom-Csv
$colCount = @($records[0].PSObject.Properties).Count
Write-Host "Source: $($records.Count) records, $colCount columns"

# Parse mappings
$mappings = $mappingsJson | ConvertFrom-Json

# Apply column mappings
$mapped = @()
foreach ($rec in $records) {{
    $newRec = [PSCustomObject]@{{}}
    foreach ($prop in $rec.PSObject.Properties) {{
        $targetCol = if ($mappings.$($prop.Name)) {{ $mappings.$($prop.Name) }} else {{ $prop.Name }}
        $newRec | Add-Member -NotePropertyName $targetCol -NotePropertyValue $prop.Value
    }}
    $mapped += $newRec
}}

$mappedCount = @($mappings.PSObject.Properties).Count
Write-Host "Mapped: $mappedCount field(s) renamed"

# Write output ({fmt})
{ps_write_block}

Write-Host "Output: $outputName"
Write-Host "Records: $($mapped.Count)"
Write-Host "Migration complete!"
'''

        # ============ PYTHON SCRIPT (Mac/Linux - python3 pre-installed) ============
        py_script = f'''#!/usr/bin/env python3
"""
Migration Script - Generated by UN Wallet
Source: {safe_filename} | Records: {len(records)} | Columns: {len(columns)} -> {len(target_columns)} | Output: {fmt}
"""
import csv, json, os, sys, base64, io

MAPPINGS = {mapping_json}
SOURCE_COLUMNS = {source_columns_json}
TARGET_COLUMNS = {target_columns_json}
SOURCE_DATA_B64 = """{source_file_b64}"""
SOURCE_FILENAME = "{source_filename_safe}"

def decode_source():
    raw = base64.b64decode(SOURCE_DATA_B64)
    ext = os.path.splitext(SOURCE_FILENAME)[1].lower()
    if ext in (".xlsx", ".xls"):
        import openpyxl; wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        ws = wb.active; rows = list(ws.iter_rows(values_only=True)); wb.close()
        if not rows: return [], []
        cols = [str(c) if c is not None else f"COL_{{i}}" for i, c in enumerate(rows[0])]
        return cols, [{{cols[i]: (str(v) if v is not None else None) for i, v in enumerate(row) if i < len(cols)}} for row in rows[1:]]
    elif ext == ".json":
        data = json.loads(raw)
        if isinstance(data, list) and len(data) > 0 and isinstance(data[0], dict):
            cols = list(data[0].keys())
            return cols, [{{k: (str(v) if v is not None else None) for k, v in item.items()}} for item in data]
        return [], []
    else:
        decoded = None
        for enc in ("utf-8-sig", "utf-8", "latin-1", "cp1252", "iso-8859-1"):
            try: decoded = raw.decode(enc); break
            except: continue
        if decoded is None: decoded = raw.decode("utf-8", errors="replace")
        import io; fio = io.StringIO(decoded)
        try: delimiter = csv.Sniffer().sniff(decoded[:1024]).delimiter
        except: delimiter = ","
        reader = csv.DictReader(fio, delimiter=delimiter)
        cols = list(reader.fieldnames) if reader.fieldnames else []
        return cols, [dict(row) for row in reader]

def migrate():
    src_cols, records = decode_source()
    if not records: print("ERROR: No records found"); sys.exit(1)
    print(f"Source: {{len(records)}} records, {{len(src_cols)}} columns")
    mapped = [{{MAPPINGS.get(k, k): v for k, v in rec.items()}} for rec in records]
    print(f"Mapped: {{len(MAPPINGS)}} field(s) renamed")
{py_write_block}
    print(f"Output: {{out}}"); print(f"Records: {{len(mapped)}}"); print("Migration complete!")

if __name__ == "__main__": migrate()
'''

        # Write both scripts
        import urllib.parse
        safe_script_name = os.path.splitext(safe_filename)[0].replace(" ", "_").replace("(", "").replace(")", "")
        ps_filename = f"migrate_{file_id}_{safe_script_name}.ps1"
        py_filename = f"migrate_{file_id}_{safe_script_name}.py"

        ps_path = os.path.join(settings.output_dir, ps_filename)
        py_path = os.path.join(settings.output_dir, py_filename)

        with open(ps_path, "w", encoding="utf-8") as f:
            f.write(ps_script)
        with open(py_path, "w", encoding="utf-8") as f:
            f.write(py_script)
        try:
            os.chmod(py_path, 0o755)
        except Exception:
            pass

        encoded_ps = urllib.parse.quote(ps_filename)
        encoded_py = urllib.parse.quote(py_filename)

        # Build the download URL the copy-paste commands point at. Prefer an
        # explicitly configured PUBLIC_BASE_URL; otherwise fall back to the host
        # the client actually reached us through (respects proxy headers).
        base_url = (settings.public_base_url or str(request.base_url)).rstrip("/")

        return {
            "success": True,
            "script_filename": py_filename,
            "download_url": f"/api/download/{encoded_py}",
            "source_columns": columns,
            "target_columns": target_columns,
            "mappings_applied": len(col_map),
            "source_file": safe_filename,
            "records_count": len(records),
            "output_format": fmt,
            "cmd_windows": f'curl -o migrate.ps1 {base_url}/api/download/{encoded_ps} && powershell -ExecutionPolicy Bypass -File migrate.ps1',
            "cmd_linux": f'curl -o migrate.py {base_url}/api/download/{encoded_py} && python3 migrate.py',
        }

    finally:
        if os.path.exists(filepath):
            try:
                os.remove(filepath)
            except OSError:
                pass


@app.get("/api/audit/{migration_id}")
@limiter.limit("30/minute")
async def get_audit(request: Request, migration_id: str, _auth=Depends(verify_api_key)):
    trail = orchestrator.get_audit_trail(migration_id)
    return {"entries": [json.loads(e.model_dump_json()) for e in trail]}


@app.get("/api/audit/{migration_id}/export")
@limiter.limit("20/minute")
async def export_audit_csv(request: Request, migration_id: str, _auth=Depends(verify_api_key)):
    import csv
    import io

    from fastapi.responses import StreamingResponse

    trail = orchestrator.get_audit_trail(migration_id)
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["timestamp", "event", "record_id", "bank_pair", "details"])
    for entry in trail:
        writer.writerow([entry.timestamp, entry.event.value, entry.record_id, entry.bank_pair, entry.details])
    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename=audit_{migration_id}.csv"},
    )


@app.post("/api/admin/cleanup")
@limiter.limit("2/minute")
async def admin_cleanup(
    request: Request,
    _auth=Depends(verify_api_key),
    target: Optional[str] = None,
    dry_run: bool = False,
):
    policy = DataRetentionPolicy(dry_run=dry_run)
    if target == "uploads":
        return {"target": "uploads", "deleted": policy.cleanup_uploads(), "dry_run": dry_run}
    elif target == "output":
        return {"target": "output", "deleted": policy.cleanup_output(), "dry_run": dry_run}
    elif target == "audit":
        return {"target": "audit", "deleted": policy.cleanup_audit_logs(), "dry_run": dry_run}
    elif target == "canonical":
        return {"target": "canonical", "deleted": policy.cleanup_canonical_store(), "dry_run": dry_run}
    else:
        report = policy.run_all()
        return {
            "target": "all",
            "dry_run": dry_run,
            "uploads_deleted": report.uploads_deleted,
            "output_deleted": report.output_deleted,
            "audit_deleted": report.audit_deleted,
            "canonical_deleted": report.canonical_deleted,
            "total_deleted": report.total_deleted,
            "errors": report.errors,
        }


# --- AI Orchestration Endpoints ---


@app.post("/api/ai/suggest-mapping")
@limiter.limit("5/minute")
async def ai_suggest_mapping(
    request: Request,
    _auth=Depends(verify_api_key),
    source_bank: str = Form(...),
    target_bank: str = Form(...),
    target_docs: str = Form(...),
):
    """
    AI analyzes target bank docs and suggests a schema mapping.
    """
    try:
        suggestion = get_schema_ai().suggest_mapping(source_bank, target_bank, target_docs)
        return {"suggestion": suggestion}
    except Exception:
        logger.exception("AI schema suggestion failed")
        raise HTTPException(status_code=500, detail="Failed to generate schema suggestion.")


@app.post("/api/ai/apply-mapping")
@limiter.limit("5/minute")
async def ai_apply_mapping(request: Request, suggestion: dict, _auth=Depends(verify_api_key)):
    """
    Validates and saves an AI suggested mapping to the registry.
    """
    try:
        path = get_schema_ai().apply_suggestion(suggestion)
        return {"status": "success", "saved_at": path}
    except Exception:
        logger.exception("AI mapping application failed")
        raise HTTPException(status_code=400, detail="Failed to apply schema mapping.")


@app.get("/api/ai/analyze-anomaly/{migration_id}")
@limiter.limit("5/minute")
async def ai_analyze_anomaly(request: Request, migration_id: str, _auth=Depends(verify_api_key)):
    """
    AI analyzes the audit trail for a specific migration to detect quality issues.
    """
    try:
        analysis = get_anomaly_ai().analyze_audit_trail(migration_id)
        return analysis
    except Exception:
        logger.exception("AI anomaly analysis failed")
        raise HTTPException(status_code=500, detail="Failed to analyze audit trail.")


def main():
    host = os.getenv("HOST", "127.0.0.1")
    port = int(os.getenv("PORT", "8000"))
    reload = os.getenv("RELOAD", "false").lower() == "true"
    uvicorn.run("api_only:app", host=host, port=port, reload=reload)


if __name__ == "__main__":
    main()
